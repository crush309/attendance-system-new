import re
from typing import List, Tuple
import xlrd
from openpyxl import load_workbook
from datetime import datetime, timedelta


def parse_attendance_file(file_path: str, rules: dict) -> dict:
    ext = file_path.rsplit(".", 1)[-1].lower()
    if ext == "xls":
        return _parse_xls(file_path, rules)
    elif ext == "xlsx":
        return _parse_xlsx(file_path, rules)
    else:
        raise ValueError(f"Unsupported file format: .{ext}")


def _normalize_time(t: str) -> str:
    """将 '8:30' 归一化为 '08:30'"""
    parts = t.split(":")
    return f"{int(parts[0]):02d}:{int(parts[1]):02d}"


def _count_valid_punches(times: List[str], rules: dict) -> Tuple[int, dict]:
    periods = rules.get("time_periods", [])
    per_period = {}
    total = 0
    norm_times = [_normalize_time(t) for t in times]
    for i, p in enumerate(periods):
        valid_in_period = sorted([t for t in norm_times if p["start"] <= t <= p["end"]])
        has_punch = 1 if valid_in_period else 0
        per_period[i] = {"count": has_punch, "earliest": valid_in_period[0] if valid_in_period else None}
        total += has_punch
    total = min(total, 3)
    return total, per_period


def _extract_times(cell_value) -> List[str]:
    if not cell_value:
        return []
    if isinstance(cell_value, (int, float)):
        val = float(cell_value)
        if 0 <= val < 1:
            total_seconds = int(val * 86400)
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            return [f"{h:02d}:{m:02d}"]
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    parts = re.split(r"[\n\r]+", text)
    times = []
    for p in parts:
        p = p.strip()
        m = re.match(r"^(\d{1,2}:\d{2})$", p)
        if m:
            times.append(m.group(1))
    return times


def _parse_xls(file_path: str, rules: dict) -> dict:
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    return _extract_data(ws.nrows, ws.ncols, lambda r, c: ws.cell_value(r, c), rules, file_path)


def _parse_xlsx(file_path: str, rules: dict) -> dict:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    return _extract_data(ws.max_row, ws.max_column, lambda r, c: ws.cell(r + 1, c + 1).value, rules, file_path)


def _extract_data(nrows, ncols, get_cell, rules, file_path=None) -> dict:
    # ---------- 1. 查找统计日期，获取起始日期 ----------
    start_date_str = None
    start_date = None
    for r in range(min(10, nrows)):
        val = get_cell(r, 0)
        if val and "统计日期" in str(val):
            # 提取 2026/06/01-2026/06/30 中的起始日期
            parts = str(val).split("统计日期:")[-1].strip()
            if "-" in parts:
                start_part = parts.split("-")[0].strip()
                # 尝试解析 YYYY/MM/DD 或 YYYY-MM-DD
                try:
                    # 替换 / 为 -
                    start_part = start_part.replace("/", "-")
                    start_date = datetime.strptime(start_part, "%Y-%m-%d").date()
                    start_date_str = start_date.strftime("%Y-%m-%d")
                    break
                except:
                    pass

    # 如果没找到，尝试从文件名或默认值（但建议提示）
    if start_date is None:
        # 若无法获取，则回退到使用当前月份的第一天（但可能不准确）
        today = datetime.now().date()
        start_date = today.replace(day=1)
        start_date_str = start_date.strftime("%Y-%m-%d")
        # 但最好提醒用户，可以打印日志，这里不处理

    # ---------- 2. 识别表头行 ----------
    header_row = -1
    for r in range(min(10, nrows)):
        val = get_cell(r, 0)
        if val and "工号" in str(val):
            header_row = r
            break
    if header_row == -1:
        raise ValueError("无法识别表头，找不到'工号'列")

    # 收集日期列（表头中从第3列（索引3）开始，值为数字）
    date_cols = []
    for c in range(3, ncols):
        hv = get_cell(header_row, c)
        if hv is not None and str(hv).strip().isdigit():
            date_cols.append(c)

    total_days = len(date_cols)
    min_punch = rules.get("min_punch_per_day", 2)
    periods = rules.get("time_periods", [])

    records = []
    for r in range(header_row + 1, nrows):
        emp_id = get_cell(r, 0)
        name = get_cell(r, 1)
        dept = get_cell(r, 2)
        if not emp_id and not name:
            continue
        emp_id = str(emp_id).strip() if emp_id else ""
        name = str(name).strip() if name else ""
        dept = str(dept).strip() if dept else ""
        if not emp_id and not name:
            continue

        attendance_days = 0
        absent_days = 0
        abnormal_days = 0
        total_punches = 0
        daily_details = []

        for c in date_cols:
            day_num = int(str(get_cell(header_row, c)).strip())
            # 生成具体日期字符串（基于起始日期 + day_num - 1 天）
            if start_date:
                current_date = start_date + timedelta(days=day_num - 1)
                date_str = current_date.strftime("%Y-%m-%d")
            else:
                date_str = f"第{day_num}天"  # 回退

            times = _extract_times(get_cell(r, c))
            if not times:
                absent_days += 1
                daily_details.append({
                    "day": day_num,
                    "date": date_str,
                    "status": "absent",
                    "earliest": None,
                    "valid_count": 0,
                    "periods": [{"label": f"{p['start']}-{p['end']}", "earliest": None} for p in periods],
                })
            else:
                valid_count, per_period = _count_valid_punches(times, rules)
                total_punches += valid_count
                overall_earliest = None
                period_info = []
                for i, p in enumerate(periods):
                    info = per_period[i]
                    period_info.append({"label": f"{p['start']}-{p['end']}", "earliest": info["earliest"]})
                    if info["earliest"] and (not overall_earliest or info["earliest"] < overall_earliest):
                        overall_earliest = info["earliest"]
                if valid_count >= min_punch:
                    attendance_days += 1
                    status = "normal"
                else:
                    abnormal_days += 1
                    status = "abnormal"
                daily_details.append({
                    "day": day_num,
                    "date": date_str,
                    "status": status,
                    "earliest": overall_earliest,
                    "valid_count": valid_count,
                    "periods": period_info,
                })

        records.append({
            "emp_id": emp_id,
            "name": name,
            "dept": dept,
            "attendance_days": attendance_days,
            "absent_days": absent_days,
            "abnormal_days": abnormal_days,
            "total_punches": total_punches,
            "total_days": total_days,
            "daily_details": daily_details,
        })

    return {"total_days": total_days, "records": records}